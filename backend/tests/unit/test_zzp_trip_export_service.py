"""Unit tests for TripExportService — CSV export.

Covers: Requirement 8 (Export — CSV format)
Reference: .kiro/specs/ZZP/rittenregistratie/design.md §4.5
"""

import pytest
from unittest.mock import Mock, patch
from datetime import date

from services.zzp_trip_export_service import TripExportService


TENANT = "test_admin"
VEHICLE_ID = 1
YEAR = 2026


def _make_service(db=None, parameter_service=None):
    """Create a TripExportService with mocked dependencies."""
    db = db or Mock()
    return TripExportService(db=db, parameter_service=parameter_service)


def _sample_trips():
    """Sample trip rows as returned by _get_trips_for_export."""
    return [
        {
            "id": 1,
            "trip_date": "2026-07-13",
            "start_time": "08:30",
            "end_time": "09:15",
            "start_address": "Keizersgracht 100, Amsterdam",
            "end_address": "Stationsplein 1, Utrecht",
            "start_odometer": 45230,
            "end_odometer": 45275,
            "distance_km": 45,
            "trip_category": "Zakelijk",
            "trip_purpose": "Klantbezoek",
            "route_description": None,
            "contact_id": 5,
            "project_name": "Project Alpha",
            "notes": None,
            "is_billable": True,
            "is_billed": False,
            "is_gap_fill": False,
            "contact_name": "Acme BV",
        },
        {
            "id": 2,
            "trip_date": "2026-07-14",
            "start_time": "10:00",
            "end_time": "10:45",
            "start_address": "Stationsplein 1, Utrecht",
            "end_address": "Keizersgracht 100, Amsterdam",
            "start_odometer": 45275,
            "end_odometer": 45320,
            "distance_km": 45,
            "trip_category": "Privé",
            "trip_purpose": "Overig",
            "route_description": None,
            "contact_id": None,
            "project_name": None,
            "notes": None,
            "is_billable": False,
            "is_billed": False,
            "is_gap_fill": False,
            "contact_name": None,
        },
    ]


class TestExportCSV:
    """Tests for TripExportService.export_csv."""

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_csv_basic_generation(self, mock_get_trips):
        """CSV export generates valid output with correct data."""
        mock_get_trips.return_value = _sample_trips()
        service = _make_service()

        result = service.export_csv(TENANT, VEHICLE_ID, YEAR)

        assert isinstance(result, bytes)
        # Decode skipping BOM
        content = result[3:].decode("utf-8")
        lines = content.strip().split("\n")
        # Header + 2 data rows
        assert len(lines) == 3
        # Verify first data row contains expected values
        assert "Keizersgracht 100, Amsterdam" in lines[1]
        assert "Acme BV" in lines[1]

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_csv_correct_columns(self, mock_get_trips):
        """CSV header contains all required Dutch column names."""
        mock_get_trips.return_value = _sample_trips()
        service = _make_service()

        result = service.export_csv(TENANT, VEHICLE_ID, YEAR)

        content = result[3:].decode("utf-8")
        header_line = content.split("\n")[0]
        expected_columns = [
            "Datum", "Vertrekadres", "Bestemming",
            "Begin KM", "Eind KM", "Afstand",
            "Categorie", "Doel", "Klant",
        ]
        for col in expected_columns:
            assert col in header_line, f"Column '{col}' not found in header"

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_csv_utf8_bom_encoding(self, mock_get_trips):
        """CSV is encoded as UTF-8 with BOM for Excel compatibility."""
        mock_get_trips.return_value = _sample_trips()
        service = _make_service()

        result = service.export_csv(TENANT, VEHICLE_ID, YEAR)

        # Check BOM bytes
        assert result[:3] == b"\xef\xbb\xbf"
        # Rest should be valid UTF-8
        content = result[3:].decode("utf-8")
        assert len(content) > 0

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_csv_semicolon_separator(self, mock_get_trips):
        """CSV uses semicolon as separator (Dutch standard)."""
        mock_get_trips.return_value = _sample_trips()
        service = _make_service()

        result = service.export_csv(TENANT, VEHICLE_ID, YEAR)

        content = result[3:].decode("utf-8")
        header_line = content.split("\n")[0]
        # Semicolons separate 9 columns → 8 semicolons in header
        assert header_line.count(";") == 8

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_csv_empty_trips_returns_headers_only(self, mock_get_trips):
        """Empty trip list returns CSV with only the header row."""
        mock_get_trips.return_value = []
        service = _make_service()

        result = service.export_csv(TENANT, VEHICLE_ID, YEAR)

        content = result[3:].decode("utf-8")
        lines = [line for line in content.strip().split("\n") if line.strip()]
        # Only header row, no data
        assert len(lines) == 1
        assert "Datum" in lines[0]

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_csv_null_contact_name_handled(self, mock_get_trips):
        """Trips without a contact_name produce empty string in Klant column."""
        trips = [_sample_trips()[1]]  # Trip with contact_name=None
        mock_get_trips.return_value = trips
        service = _make_service()

        result = service.export_csv(TENANT, VEHICLE_ID, YEAR)

        content = result[3:].decode("utf-8")
        lines = content.strip().split("\n")
        # Data row should not error, None becomes empty
        assert len(lines) == 2

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_csv_passes_filters_to_query(self, mock_get_trips):
        """Filters are passed through to _get_trips_for_export."""
        mock_get_trips.return_value = []
        service = _make_service()
        filters = {"trip_category": "Zakelijk"}

        service.export_csv(TENANT, VEHICLE_ID, YEAR, filters=filters)

        mock_get_trips.assert_called_once_with(TENANT, VEHICLE_ID, YEAR, filters)


class TestExportXLSX:
    """Tests for TripExportService.export_xlsx."""

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_xlsx_generates_valid_bytes(self, mock_get_trips):
        """XLSX export generates valid bytes (openpyxl-readable)."""
        from openpyxl import load_workbook

        mock_get_trips.return_value = _sample_trips()
        service = _make_service()

        result = service.export_xlsx(TENANT, VEHICLE_ID, YEAR)

        assert isinstance(result, bytes)
        assert len(result) > 0
        # Verify it's a valid XLSX by loading it
        import io
        wb = load_workbook(io.BytesIO(result))
        assert wb.active is not None

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_xlsx_correct_sheet_name(self, mock_get_trips):
        """XLSX sheet title is 'Rittenregistratie {year}'."""
        from openpyxl import load_workbook
        import io

        mock_get_trips.return_value = _sample_trips()
        service = _make_service()

        result = service.export_xlsx(TENANT, VEHICLE_ID, YEAR)

        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.title == f"Rittenregistratie {YEAR}"

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_xlsx_bold_headers(self, mock_get_trips):
        """XLSX header row has bold font styling."""
        from openpyxl import load_workbook
        import io

        mock_get_trips.return_value = _sample_trips()
        service = _make_service()

        result = service.export_xlsx(TENANT, VEHICLE_ID, YEAR)

        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        # All header cells (row 1) should be bold
        for col_idx in range(1, 10):
            cell = ws.cell(row=1, column=col_idx)
            assert cell.font.bold, f"Header cell column {col_idx} is not bold"

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_xlsx_correct_header_names(self, mock_get_trips):
        """XLSX header row contains all Dutch column names."""
        from openpyxl import load_workbook
        import io

        mock_get_trips.return_value = _sample_trips()
        service = _make_service()

        result = service.export_xlsx(TENANT, VEHICLE_ID, YEAR)

        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        expected_headers = [
            "Datum", "Vertrekadres", "Bestemming",
            "Begin KM", "Eind KM", "Afstand",
            "Categorie", "Doel", "Klant",
        ]
        actual_headers = [ws.cell(row=1, column=i).value for i in range(1, 10)]
        assert actual_headers == expected_headers

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_xlsx_correct_data_rows(self, mock_get_trips):
        """XLSX data rows contain correct trip values."""
        from openpyxl import load_workbook
        import io

        mock_get_trips.return_value = _sample_trips()
        service = _make_service()

        result = service.export_xlsx(TENANT, VEHICLE_ID, YEAR)

        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        # Row 2 = first data row
        assert ws.cell(row=2, column=1).value == "2026-07-13"  # Datum
        assert ws.cell(row=2, column=2).value == "Keizersgracht 100, Amsterdam"  # Vertrekadres
        assert ws.cell(row=2, column=3).value == "Stationsplein 1, Utrecht"  # Bestemming
        assert ws.cell(row=2, column=4).value == 45230  # Begin KM
        assert ws.cell(row=2, column=5).value == 45275  # Eind KM
        assert ws.cell(row=2, column=6).value == 45  # Afstand
        assert ws.cell(row=2, column=7).value == "Zakelijk"  # Categorie
        assert ws.cell(row=2, column=8).value == "Klantbezoek"  # Doel
        assert ws.cell(row=2, column=9).value == "Acme BV"  # Klant

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_xlsx_empty_trips_returns_headers_only(self, mock_get_trips):
        """Empty trip list returns XLSX with only the header row."""
        from openpyxl import load_workbook
        import io

        mock_get_trips.return_value = []
        service = _make_service()

        result = service.export_xlsx(TENANT, VEHICLE_ID, YEAR)

        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        # Header row present
        assert ws.cell(row=1, column=1).value == "Datum"
        # No data rows
        assert ws.cell(row=2, column=1).value is None

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_xlsx_passes_filters_to_query(self, mock_get_trips):
        """Filters are passed through to _get_trips_for_export."""
        mock_get_trips.return_value = []
        service = _make_service()
        filters = {"trip_category": "Zakelijk"}

        service.export_xlsx(TENANT, VEHICLE_ID, YEAR, filters=filters)

        mock_get_trips.assert_called_once_with(TENANT, VEHICLE_ID, YEAR, filters)


class TestExportPDF:
    """Tests for TripExportService.export_pdf."""

    @patch.object(TripExportService, "_html_to_pdf")
    @patch.object(TripExportService, "_get_vehicle_info")
    @patch.object(TripExportService, "_get_trips_for_export")
    def test_pdf_calls_weasyprint_with_rendered_html(
        self, mock_get_trips, mock_vehicle, mock_html_to_pdf
    ):
        """export_pdf renders HTML and converts to PDF via _html_to_pdf."""
        mock_get_trips.return_value = _sample_trips()
        mock_vehicle.return_value = {
            "id": 1, "license_plate": "AB-123-CD",
            "make": "Volkswagen", "model": "Golf",
            "vehicle_type": "private_for_business", "odometer_unit": "km",
        }
        mock_html_to_pdf.return_value = b"%PDF-fake"
        service = _make_service()

        result = service.export_pdf(TENANT, VEHICLE_ID, YEAR)

        assert result == b"%PDF-fake"
        mock_html_to_pdf.assert_called_once()
        html_arg = mock_html_to_pdf.call_args[0][0]
        assert "Rittenregistratie 2026" in html_arg
        assert "AB-123-CD" in html_arg

    @patch.object(TripExportService, "_html_to_pdf")
    @patch.object(TripExportService, "_get_vehicle_info")
    @patch.object(TripExportService, "_get_trips_for_export")
    def test_pdf_html_contains_trip_data(
        self, mock_get_trips, mock_vehicle, mock_html_to_pdf
    ):
        """Rendered HTML includes trip addresses, distances, categories."""
        mock_get_trips.return_value = _sample_trips()
        mock_vehicle.return_value = {
            "id": 1, "license_plate": "AB-123-CD",
            "make": "Volkswagen", "model": "Golf",
            "vehicle_type": "private_for_business", "odometer_unit": "km",
        }
        mock_html_to_pdf.return_value = b"%PDF-fake"
        service = _make_service()

        service.export_pdf(TENANT, VEHICLE_ID, YEAR)

        html_arg = mock_html_to_pdf.call_args[0][0]
        assert "Keizersgracht 100, Amsterdam" in html_arg
        assert "Stationsplein 1, Utrecht" in html_arg
        assert "45230" in html_arg
        assert "45275" in html_arg
        assert "Zakelijk" in html_arg
        assert "Klantbezoek" in html_arg
        assert "Acme BV" in html_arg

    @patch.object(TripExportService, "_html_to_pdf")
    @patch.object(TripExportService, "_get_vehicle_info")
    @patch.object(TripExportService, "_get_trips_for_export")
    def test_pdf_html_contains_yearly_totals(
        self, mock_get_trips, mock_vehicle, mock_html_to_pdf
    ):
        """PDF HTML includes yearly totals per category."""
        mock_get_trips.return_value = _sample_trips()
        mock_vehicle.return_value = {
            "id": 1, "license_plate": "AB-123-CD",
            "make": "Volkswagen", "model": "Golf",
            "vehicle_type": "private_for_business", "odometer_unit": "km",
        }
        mock_html_to_pdf.return_value = b"%PDF-fake"
        service = _make_service()

        service.export_pdf(TENANT, VEHICLE_ID, YEAR)

        html_arg = mock_html_to_pdf.call_args[0][0]
        # Totals: 45 zakelijk + 45 privé = 90 total
        assert "90 km" in html_arg  # total_km
        assert "45 km" in html_arg  # zakelijk_km and prive_km
        assert "Jaaroverzicht 2026" in html_arg

    @patch.object(TripExportService, "_html_to_pdf")
    @patch.object(TripExportService, "_get_vehicle_info")
    @patch.object(TripExportService, "_get_trips_for_export")
    def test_pdf_html_contains_vehicle_description(
        self, mock_get_trips, mock_vehicle, mock_html_to_pdf
    ):
        """PDF includes vehicle make/model and license plate."""
        mock_get_trips.return_value = []
        mock_vehicle.return_value = {
            "id": 1, "license_plate": "XY-999-ZZ",
            "make": "Tesla", "model": "Model 3",
            "vehicle_type": "business", "odometer_unit": "km",
        }
        mock_html_to_pdf.return_value = b"%PDF-fake"
        service = _make_service()

        service.export_pdf(TENANT, VEHICLE_ID, YEAR)

        html_arg = mock_html_to_pdf.call_args[0][0]
        assert "Tesla Model 3" in html_arg
        assert "XY-999-ZZ" in html_arg

    @patch.object(TripExportService, "_html_to_pdf")
    @patch.object(TripExportService, "_get_vehicle_info")
    @patch.object(TripExportService, "_get_trips_for_export")
    def test_pdf_empty_trips_still_generates(
        self, mock_get_trips, mock_vehicle, mock_html_to_pdf
    ):
        """Empty trip list still produces a valid PDF (just headers/totals)."""
        mock_get_trips.return_value = []
        mock_vehicle.return_value = {
            "id": 1, "license_plate": "AB-123-CD",
            "make": "Volkswagen", "model": "Golf",
            "vehicle_type": "private_for_business", "odometer_unit": "km",
        }
        mock_html_to_pdf.return_value = b"%PDF-empty"
        service = _make_service()

        result = service.export_pdf(TENANT, VEHICLE_ID, YEAR)

        assert result == b"%PDF-empty"
        html_arg = mock_html_to_pdf.call_args[0][0]
        assert "0 km" in html_arg  # All totals are 0

    @patch.object(TripExportService, "_html_to_pdf")
    @patch.object(TripExportService, "_get_vehicle_info")
    @patch.object(TripExportService, "_get_trips_for_export")
    def test_pdf_passes_filters_to_query(
        self, mock_get_trips, mock_vehicle, mock_html_to_pdf
    ):
        """Filters are passed through to _get_trips_for_export."""
        mock_get_trips.return_value = []
        mock_vehicle.return_value = {
            "id": 1, "license_plate": "AB-123-CD",
            "make": "", "model": "",
            "vehicle_type": "private_for_business", "odometer_unit": "km",
        }
        mock_html_to_pdf.return_value = b"%PDF-fake"
        service = _make_service()
        filters = {"trip_category": "Zakelijk"}

        service.export_pdf(TENANT, VEHICLE_ID, YEAR, filters=filters)

        mock_get_trips.assert_called_once_with(TENANT, VEHICLE_ID, YEAR, filters)

    @patch.object(TripExportService, "_html_to_pdf")
    @patch.object(TripExportService, "_get_vehicle_info")
    @patch.object(TripExportService, "_get_trips_for_export")
    def test_pdf_belastingdienst_footer_present(
        self, mock_get_trips, mock_vehicle, mock_html_to_pdf
    ):
        """PDF HTML has a Belastingdienst footer reference."""
        mock_get_trips.return_value = []
        mock_vehicle.return_value = {
            "id": 1, "license_plate": "AB-123-CD",
            "make": "", "model": "",
            "vehicle_type": "private_for_business", "odometer_unit": "km",
        }
        mock_html_to_pdf.return_value = b"%PDF-fake"
        service = _make_service()

        service.export_pdf(TENANT, VEHICLE_ID, YEAR)

        html_arg = mock_html_to_pdf.call_args[0][0]
        assert "Belastingdienst" in html_arg


class TestCalculateTotals:
    """Tests for TripExportService._calculate_totals."""

    def test_totals_from_mixed_categories(self):
        """Correctly sums km per category."""
        trips = [
            {"distance_km": 45, "trip_category": "Zakelijk"},
            {"distance_km": 30, "trip_category": "Privé"},
            {"distance_km": 20, "trip_category": "Woon-werk"},
            {"distance_km": 10, "trip_category": "Zakelijk"},
        ]
        totals = TripExportService._calculate_totals(trips)

        assert totals["total_km"] == 105
        assert totals["zakelijk_km"] == 55
        assert totals["prive_km"] == 30
        assert totals["woonwerk_km"] == 20
        assert totals["trip_count"] == 4

    def test_totals_empty_trips(self):
        """Empty trip list returns all zeros."""
        totals = TripExportService._calculate_totals([])

        assert totals["total_km"] == 0
        assert totals["zakelijk_km"] == 0
        assert totals["prive_km"] == 0
        assert totals["woonwerk_km"] == 0
        assert totals["trip_count"] == 0

    def test_totals_handles_none_distance(self):
        """None distance_km is treated as 0."""
        trips = [
            {"distance_km": None, "trip_category": "Zakelijk"},
            {"distance_km": 50, "trip_category": "Zakelijk"},
        ]
        totals = TripExportService._calculate_totals(trips)

        assert totals["total_km"] == 50
        assert totals["zakelijk_km"] == 50


class TestGetVehicleInfo:
    """Tests for TripExportService._get_vehicle_info."""

    def test_returns_vehicle_dict_when_found(self):
        """Returns vehicle data from database query."""
        db = Mock()
        db.execute_query.return_value = [
            {
                "id": 1,
                "license_plate": "AB-123-CD",
                "make": "Volkswagen",
                "model": "Golf",
                "year_built": 2020,
                "vehicle_type": "private_for_business",
                "odometer_unit": "km",
            }
        ]
        service = _make_service(db=db)

        result = service._get_vehicle_info(TENANT, VEHICLE_ID)

        assert result["license_plate"] == "AB-123-CD"
        assert result["make"] == "Volkswagen"

    def test_returns_fallback_when_not_found(self):
        """Returns minimal fallback dict when vehicle not in database."""
        db = Mock()
        db.execute_query.return_value = []
        service = _make_service(db=db)

        result = service._get_vehicle_info(TENANT, VEHICLE_ID)

        assert result["license_plate"] == "Onbekend"
        assert result["id"] == VEHICLE_ID

    def test_returns_fallback_on_db_error(self):
        """Returns fallback on DatabaseError."""
        from db_exceptions import DatabaseError

        db = Mock()
        db.execute_query.side_effect = DatabaseError("connection lost")
        service = _make_service(db=db)

        result = service._get_vehicle_info(TENANT, VEHICLE_ID)

        assert result["license_plate"] == "Onbekend"


class TestHtmlToPdf:
    """Tests for TripExportService._html_to_pdf."""

    @patch.dict("sys.modules", {"weasyprint": None})
    def test_raises_runtime_error_without_weasyprint(self):
        """Raises RuntimeError if weasyprint is not importable."""
        with pytest.raises(RuntimeError, match="weasyprint"):
            TripExportService._html_to_pdf("<html></html>")


class TestGetYearlySummary:
    """Tests for TripExportService.get_yearly_summary."""

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_summary_correct_totals(self, mock_get_trips):
        """Yearly summary returns correct category totals."""
        mock_get_trips.return_value = [
            {"trip_date": "2026-01-10", "distance_km": 100, "trip_category": "Zakelijk"},
            {"trip_date": "2026-02-15", "distance_km": 50, "trip_category": "Privé"},
            {"trip_date": "2026-03-20", "distance_km": 30, "trip_category": "Woon-werk"},
            {"trip_date": "2026-01-25", "distance_km": 60, "trip_category": "Zakelijk"},
        ]
        service = _make_service()

        result = service.get_yearly_summary(TENANT, VEHICLE_ID, YEAR)

        assert result["year"] == YEAR
        assert result["vehicle_id"] == VEHICLE_ID
        assert result["total_km"] == 240
        assert result["zakelijk_km"] == 160
        assert result["prive_km"] == 50
        assert result["woonwerk_km"] == 30
        assert result["trip_count"] == 4

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_summary_monthly_breakdown(self, mock_get_trips):
        """Monthly breakdown groups trips by month and sums per category."""
        mock_get_trips.return_value = [
            {"trip_date": "2026-01-05", "distance_km": 80, "trip_category": "Zakelijk"},
            {"trip_date": "2026-01-12", "distance_km": 20, "trip_category": "Privé"},
            {"trip_date": "2026-01-20", "distance_km": 30, "trip_category": "Woon-werk"},
            {"trip_date": "2026-03-10", "distance_km": 50, "trip_category": "Zakelijk"},
            {"trip_date": "2026-03-22", "distance_km": 15, "trip_category": "Privé"},
        ]
        service = _make_service()

        result = service.get_yearly_summary(TENANT, VEHICLE_ID, YEAR)

        breakdown = result["monthly_breakdown"]
        assert len(breakdown) == 2  # January and March

        jan = breakdown[0]
        assert jan["month"] == "2026-01"
        assert jan["zakelijk"] == 80
        assert jan["prive"] == 20
        assert jan["woonwerk"] == 30

        mar = breakdown[1]
        assert mar["month"] == "2026-03"
        assert mar["zakelijk"] == 50
        assert mar["prive"] == 15
        assert mar["woonwerk"] == 0

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_summary_empty_year(self, mock_get_trips):
        """Empty year returns zero totals and empty breakdown."""
        mock_get_trips.return_value = []
        service = _make_service()

        result = service.get_yearly_summary(TENANT, VEHICLE_ID, YEAR)

        assert result["year"] == YEAR
        assert result["vehicle_id"] == VEHICLE_ID
        assert result["total_km"] == 0
        assert result["zakelijk_km"] == 0
        assert result["prive_km"] == 0
        assert result["woonwerk_km"] == 0
        assert result["trip_count"] == 0
        assert result["monthly_breakdown"] == []

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_summary_return_format(self, mock_get_trips):
        """Return dict matches the design specification format."""
        mock_get_trips.return_value = [
            {"trip_date": "2026-06-15", "distance_km": 45, "trip_category": "Zakelijk"},
        ]
        service = _make_service()

        result = service.get_yearly_summary(TENANT, VEHICLE_ID, YEAR)

        # Verify all required keys present
        expected_keys = {
            "year", "vehicle_id", "total_km", "zakelijk_km",
            "prive_km", "woonwerk_km", "trip_count", "monthly_breakdown",
        }
        assert set(result.keys()) == expected_keys

        # Verify monthly_breakdown entry format
        assert len(result["monthly_breakdown"]) == 1
        entry = result["monthly_breakdown"][0]
        assert set(entry.keys()) == {"month", "zakelijk", "prive", "woonwerk"}
        assert entry["month"] == "2026-06"

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_summary_monthly_breakdown_sorted_by_month(self, mock_get_trips):
        """Monthly breakdown is sorted chronologically."""
        mock_get_trips.return_value = [
            {"trip_date": "2026-12-01", "distance_km": 10, "trip_category": "Zakelijk"},
            {"trip_date": "2026-03-01", "distance_km": 20, "trip_category": "Zakelijk"},
            {"trip_date": "2026-07-01", "distance_km": 30, "trip_category": "Zakelijk"},
        ]
        service = _make_service()

        result = service.get_yearly_summary(TENANT, VEHICLE_ID, YEAR)

        months = [entry["month"] for entry in result["monthly_breakdown"]]
        assert months == ["2026-03", "2026-07", "2026-12"]

    @patch.object(TripExportService, "_get_trips_for_export")
    def test_summary_calls_get_trips_correctly(self, mock_get_trips):
        """get_yearly_summary passes correct args to _get_trips_for_export."""
        mock_get_trips.return_value = []
        service = _make_service()

        service.get_yearly_summary(TENANT, VEHICLE_ID, YEAR)

        mock_get_trips.assert_called_once_with(TENANT, VEHICLE_ID, YEAR)

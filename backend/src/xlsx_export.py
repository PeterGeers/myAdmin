"""XLSX Export Processor - core export infrastructure.

Provides the XLSXExportProcessor class for generating Excel workbooks
from financial ledger data and downloading associated document files.

Styling is delegated to xlsx_styles module.
File-download helpers and progress-aware export methods are provided
by the XLSXProgressExportMixin in xlsx_report_generators.
"""

import os
import shutil
from database import DatabaseManager
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from services.template_service import TemplateService
from xlsx_styles import apply_worksheet_formatting
from xlsx_report_generators import XLSXProgressExportMixin
import logging

logger = logging.getLogger(__name__)


class XLSXExportProcessor(XLSXProgressExportMixin):
    """Processes financial ledger data into Excel workbooks with associated files.
    
    Inherits file-download helpers and progress-reporting export methods
    from XLSXProgressExportMixin.
    """

    def __init__(self, test_mode=False, parameter_service=None, tenant: str = None):
        self.test_mode = test_mode
        self.db = DatabaseManager(test_mode=test_mode)
        self.template_service = TemplateService(self.db)
        # Default template path (fallback)
        self.default_template_path = os.path.join(
            os.path.dirname(__file__), '..', 'templates', 'xlsx', 'template.xlsx'
        )
        # Output path from parameter service or environment-based default
        if parameter_service and tenant:
            self.default_output_base_path = parameter_service.get_param(
                'storage', 'report_output_path', tenant=tenant
            )
        else:
            self.default_output_base_path = None
        if not self.default_output_base_path:
            if os.getenv('DOCKER_ENV') or os.path.exists('/.dockerenv'):
                self.default_output_base_path = '/app/reports'
            else:
                self.default_output_base_path = os.path.join(os.getcwd(), 'reports')
        os.makedirs(self.default_output_base_path, exist_ok=True)
        self.folder_search_log = []
        logger.info("XLSXExportProcessor initialized with TemplateService")

    def _get_template_path(self, administration):
        """Get template path for administration using TemplateService."""
        try:
            metadata = self.template_service.get_template_metadata(
                administration, 'financial_report_xlsx'
            )
            if metadata and metadata.get('field_mappings'):
                field_mappings = metadata['field_mappings']
                template_path = field_mappings.get('template_path')
                if template_path and os.path.exists(template_path):
                    logger.info(
                        f"Using template path from TemplateService for {administration}: "
                        f"{template_path}"
                    )
                    return template_path
                else:
                    logger.warning(
                        f"Template path from database not found: {template_path}, "
                        f"using default"
                    )
        except Exception as e:
            logger.warning(f"Could not get template path from TemplateService: {e}")

        logger.info(f"Using default template path for {administration}")
        return self.default_template_path

    def _get_output_base_path(self, administration):
        """Get output base path for administration using TemplateService."""
        try:
            metadata = self.template_service.get_template_metadata(
                administration, 'financial_report_xlsx'
            )
            if metadata and metadata.get('field_mappings'):
                field_mappings = metadata['field_mappings']
                output_path = field_mappings.get('output_base_path')
                if output_path:
                    os.makedirs(output_path, exist_ok=True)
                    logger.info(
                        f"Using output path from TemplateService for {administration}: "
                        f"{output_path}"
                    )
                    return output_path
        except Exception as e:
            logger.warning(f"Could not get output path from TemplateService: {e}")

        logger.info(f"Using default output path for {administration}")
        return self.default_output_base_path

    def make_ledgers(self, year, administration):
        """Calculate starting balance and add all transactions for a specific year/administration."""
        conn = self.db.get_connection()
        cursor = conn.cursor(dictionary=True)

        try:
            # Get balance accounts (VW = N) for years before target year
            balance_query = """
                SELECT Reknum, AccountName, Parent, Administration,
                       SUM(Amount) as Amount
                FROM vw_mutaties 
                WHERE VW = 'N' AND Administration LIKE %s AND jaar < %s
                GROUP BY Reknum, AccountName, Parent, Administration
                HAVING ABS(SUM(Amount)) > 0.01
            """
            cursor.execute(balance_query, [f"{administration}%", year])
            balance_data = cursor.fetchall()

            # Create beginning balance records
            beginning_balance = []
            for row in balance_data:
                balance_record = {
                    'TransactionNumber': f'Beginbalans {year}',
                    'TransactionDate': f'{year}-01-01',
                    'TransactionDescription': f'Beginbalans van het jaar {year} van Administratie {administration}',
                    'Amount': round(row['Amount'], 2),
                    'Reknum': row['Reknum'],
                    'AccountName': row['AccountName'],
                    'Parent': row['Parent'],
                    'Administration': row['Administration'],
                    'VW': 'N',
                    'jaar': year,
                    'kwartaal': 1,
                    'maand': 1,
                    'week': 1,
                    'ReferenceNumber': '',
                    'DocUrl': '',
                    'Document': ''
                }
                beginning_balance.append(balance_record)

            # Get all transactions for the specific year
            transactions_query = """
                SELECT TransactionNumber, TransactionDate, TransactionDescription, Amount, 
                       Reknum, AccountName, Parent, Administration, VW, jaar, kwartaal, 
                       maand, week, ReferenceNumber, Ref3 as DocUrl, Ref4 as Document
                FROM vw_mutaties 
                WHERE Administration LIKE %s AND jaar = %s
                ORDER BY TransactionDate, Reknum
            """
            cursor.execute(transactions_query, [f"{administration}%", year])
            transactions_data = cursor.fetchall()

            # Combine beginning balance and transactions
            all_data = beginning_balance + transactions_data
            return all_data

        finally:
            cursor.close()
            conn.close()

    def write_workbook(self, data, filename, sheet_name='data', administration=None):
        """Write data to Excel workbook using template."""
        # Ensure output directory exists
        output_dir = os.path.dirname(filename)
        if output_dir:
            try:
                os.makedirs(output_dir, exist_ok=True)
            except OSError as e:
                print(f"Warning: Could not create directory {output_dir}: {e}")
                filename = os.path.basename(filename)
                print(f"Falling back to current directory: {filename}")

        # Get template path for this administration
        template_path = self._get_template_path(administration) if administration else self.default_template_path

        # Copy template if target file doesn't exist
        if not os.path.exists(filename):
            if os.path.exists(template_path):
                shutil.copy2(template_path, filename)
            else:
                from openpyxl import Workbook
                wb = Workbook()
                wb.save(filename)

        # Load workbook
        wb = load_workbook(filename)

        # Remove existing data sheet if it exists
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

        # Create new data sheet
        ws = wb.create_sheet(sheet_name)

        # Convert data to DataFrame and select required columns
        df = pd.DataFrame(data)
        required_columns = [
            'TransactionNumber', 'TransactionDate', 'TransactionDescription',
            'Amount', 'Reknum', 'AccountName', 'SubParent', 'Parent', 'VW',
            'jaar', 'kwartaal', 'maand', 'week', 'ReferenceNumber',
            'Administration', 'DocUrl', 'Document'
        ]
        available_columns = [col for col in required_columns if col in df.columns]
        df = df[available_columns]

        # Convert Amount column to numeric
        if 'Amount' in df.columns:
            df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')

        # Clean string columns (but not Amount)
        for col in df.select_dtypes(include=['object']).columns:
            if col != 'Amount':
                df[col] = df[col].astype(str).str.strip()

        # Write data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Apply formatting from xlsx_styles
        apply_worksheet_formatting(ws)

        # Save workbook
        wb.save(filename)
        return filename

    def generate_xlsx_export(self, administrations, years):
        """Generate XLSX exports for specified administrations and years."""
        results = []

        for administration in administrations:
            for year in years:
                try:
                    print(f"Processing {administration} {year}")
                    ledger_data = self.make_ledgers(year, administration)
                    print(f"Found {len(ledger_data)} records")

                    if not ledger_data:
                        results.append({
                            'administration': administration,
                            'year': year,
                            'error': 'No data found for this administration/year combination',
                            'success': False
                        })
                        continue

                    output_base_path = self._get_output_base_path(administration)
                    folder_path = os.path.join(output_base_path, f"{administration}{year}")
                    filename = os.path.join(folder_path, f"{administration}{year}.xlsx")

                    try:
                        os.makedirs(output_base_path, exist_ok=True)
                    except OSError as e:
                        print(f"Warning: Could not create base output directory {output_base_path}: {e}")
                        filename = f"{administration}{year}.xlsx"
                        print(f"Using current directory for output: {filename}")

                    output_file = self.write_workbook(ledger_data, filename, 'data', administration)
                    print(f"Created file: {output_file}")

                    file_count = self.export_files(ledger_data, year, administration)

                    results.append({
                        'administration': administration,
                        'year': year,
                        'filename': output_file,
                        'records': len(ledger_data),
                        'files_processed': file_count,
                        'success': True
                    })

                except Exception as e:
                    print(f"Error processing {administration} {year}: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    results.append({
                        'administration': administration,
                        'year': year,
                        'error': str(e),
                        'success': False
                    })

        return results

"""
Chart of Accounts Management Routes for myAdmin

This module provides API endpoints for managing the chart of accounts (rekeningschema).
Requires Tenant_Admin role and FIN module access.

Endpoints:
- GET    /api/tenant-admin/chart-of-accounts           - List all accounts
- GET    /api/tenant-admin/chart-of-accounts/<account> - Get single account
- POST   /api/tenant-admin/chart-of-accounts           - Create account
- PUT    /api/tenant-admin/chart-of-accounts/<account> - Update account
- DELETE /api/tenant-admin/chart-of-accounts/<account> - Delete account
- GET    /api/tenant-admin/chart-of-accounts/export    - Export to Excel
- POST   /api/tenant-admin/chart-of-accounts/import    - Import from Excel

Based on spec: .kiro/specs/FIN/Chart of Accounts Management/
"""

from flask import Blueprint, request, jsonify, send_file
from auth.cognito_utils import cognito_required
from auth.tenant_context import get_current_tenant, get_user_tenants, is_tenant_admin
from database import DatabaseManager
import os
import logging
from typing import Dict, Any
from datetime import datetime
from io import BytesIO

# Initialize logger
logger = logging.getLogger(__name__)

# Create blueprint
chart_of_accounts_bp = Blueprint('chart_of_accounts', __name__)


# ============================================================================
# Helper Functions
# ============================================================================


def has_fin_module(tenant: str) -> bool:
    """
    Check if a tenant has the FIN module enabled.
    
    Args:
        tenant (str): The tenant administration name
    
    Returns:
        bool: True if tenant has FIN module and it's active
    """
    try:
        test_mode = os.getenv('TEST_MODE', 'false').lower() == 'true'
        db = DatabaseManager(test_mode=test_mode)
        
        query = """
            SELECT is_active 
            FROM tenant_modules 
            WHERE administration = %s AND module_name = 'FIN'
        """
        result = db.execute_query(query, (tenant,))
        
        return bool(result and result[0].get('is_active'))
        
    except Exception as e:
        logger.error(f"Error checking FIN module for tenant {tenant}: {e}")
        return False


def validate_account_number(account: str) -> str:
    """Validate and clean an account number."""
    if account is None:
        raise ValueError("Account number is required")
    
    cleaned = account.strip()
    
    if not cleaned:
        raise ValueError("Account number is required")
    
    return cleaned


def validate_account_name(name: str) -> str:
    """Validate and clean an account name."""
    if name is None:
        raise ValueError("Account name is required")
    
    cleaned = name.strip()
    
    if not cleaned:
        raise ValueError("Account name is required")
    
    return cleaned


def is_account_used_in_transactions(tenant: str, account: str) -> int:
    """
    Check if an account is used in any transactions.
    
    Returns:
        int: Count of transactions using this account
    """
    try:
        test_mode = os.getenv('TEST_MODE', 'false').lower() == 'true'
        db = DatabaseManager(test_mode=test_mode)
        
        query = """
            SELECT COUNT(*) as count
            FROM mutaties
            WHERE administration = %s
            AND (Debet = %s OR Credit = %s)
        """
        result = db.execute_query(query, (tenant, account, account))
        
        return result[0].get('count', 0) if result else 0
        
    except Exception as e:
        logger.error(f"Error checking account usage for {account}: {e}")
        return 0


# ============================================================================
# API Endpoints
# ============================================================================


@chart_of_accounts_bp.route('/api/tenant-admin/chart-of-accounts', methods=['GET'])
@cognito_required(required_permissions=[])
def list_accounts(user_email, user_roles):
    """
    Get all accounts for current tenant with search, sorting, and pagination.
    
    Query Parameters:
        - search: string (optional) - Search in account number or name
        - sort_by: string (optional) - Column to sort by
        - sort_order: string (optional) - 'asc' or 'desc'
        - page: int (optional) - Page number (default: 1)
        - limit: int (optional) - Items per page (default: 50)
    
    Returns:
        {
            "success": true,
            "accounts": [...],
            "total": 150,
            "page": 1,
            "limit": 50,
            "pages": 3
        }
    """
    try:
        # Get tenant from request
        tenant = get_current_tenant(request)
        
        # Extract user tenants from JWT
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            jwt_token = auth_header.replace('Bearer ', '').strip()
            user_tenants = get_user_tenants(jwt_token)
        else:
            return jsonify({'error': 'Invalid authorization'}), 401
        
        # Check if user is tenant admin
        if not is_tenant_admin(user_roles, tenant, user_tenants):
            return jsonify({'error': 'Tenant admin access required'}), 403
        
        # Check FIN module access
        if not has_fin_module(tenant):
            return jsonify({'error': 'FIN module not enabled'}), 403
        
        # Get query parameters
        search = request.args.get('search', '').strip()
        sort_by = request.args.get('sort_by', 'Account')
        sort_order = request.args.get('sort_order', 'asc').lower()
        
        try:
            page = int(request.args.get('page', 1))
            limit = int(request.args.get('limit', 50))
        except ValueError:
            return jsonify({'error': 'Invalid page or limit parameter'}), 400
        
        # Validate page and limit
        if page < 1:
            page = 1
        if limit < 1 or limit > 1000:
            limit = 50
        
        # Initialize database
        test_mode = os.getenv('TEST_MODE', 'false').lower() == 'true'
        db = DatabaseManager(test_mode=test_mode)
        
        # Build query
        query = """
            SELECT AccountID, Account, AccountName, AccountLookup, 
                   SubParent, Parent, VW, Belastingaangifte, 
                   administration, Pattern,
                   JSON_UNQUOTE(JSON_EXTRACT(parameters, '$.purpose')) as purpose
            FROM rekeningschema
            WHERE administration = %s
        """
        params = [tenant]
        
        # Add search filter
        if search:
            query += " AND (Account LIKE %s OR AccountName LIKE %s)"
            search_pattern = f"%{search}%"
            params.extend([search_pattern, search_pattern])
        
        # Add sorting
        valid_columns = ['Account', 'AccountName', 'AccountLookup', 'Belastingaangifte', 
                        'SubParent', 'Parent', 'VW', 'Pattern']
        if sort_by in valid_columns:
            order = 'ASC' if sort_order == 'asc' else 'DESC'
            query += f" ORDER BY {sort_by} {order}"
        else:
            # Default sorting
            query += " ORDER BY Account ASC"
        
        # Add pagination
        offset = (page - 1) * limit
        query += " LIMIT %s OFFSET %s"
        params.extend([limit, offset])
        
        # Execute query
        accounts = db.execute_query(query, params)
        
        # Get total count for pagination
        count_query = """
            SELECT COUNT(*) as total
            FROM rekeningschema
            WHERE administration = %s
        """
        count_params = [tenant]
        
        if search:
            count_query += " AND (Account LIKE %s OR AccountName LIKE %s)"
            count_params.extend([search_pattern, search_pattern])
        
        count_result = db.execute_query(count_query, count_params)
        total = count_result[0]['total'] if count_result else 0
        
        # Calculate total pages
        pages = (total + limit - 1) // limit if total > 0 else 0
        
        return jsonify({
            'success': True,
            'accounts': accounts,
            'total': total,
            'page': page,
            'limit': limit,
            'pages': pages
        })
        
    except Exception as e:
        logger.error(f"Error listing accounts for tenant {tenant}: {e}")
        return jsonify({'error': 'Failed to list accounts', 'details': str(e)}), 500



@chart_of_accounts_bp.route('/api/tenant-admin/chart-of-accounts/<account>', methods=['GET'])
@cognito_required(required_permissions=[])
def get_account(user_email, user_roles, account):
    """
    Get single account details.
    
    Path Parameters:
        - account: string (required) - Account number
    
    Returns:
        {
            "success": true,
            "account": {
                "Account": "1000",
                "AccountName": "Kas",
                "AccountLookup": "CASH",
                "Belastingaangifte": "Activa"
            }
        }
    """
    try:
        # Get tenant from request
        tenant = get_current_tenant(request)
        
        # Extract user tenants from JWT
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            jwt_token = auth_header.replace('Bearer ', '').strip()
            user_tenants = get_user_tenants(jwt_token)
        else:
            return jsonify({'error': 'Invalid authorization'}), 401
        
        # Check if user is tenant admin
        if not is_tenant_admin(user_roles, tenant, user_tenants):
            return jsonify({'error': 'Tenant admin access required'}), 403
        
        # Check FIN module access
        if not has_fin_module(tenant):
            return jsonify({'error': 'FIN module not enabled'}), 403
        
        # Initialize database
        test_mode = os.getenv('TEST_MODE', 'false').lower() == 'true'
        db = DatabaseManager(test_mode=test_mode)
        
        # Query account
        query = """
            SELECT AccountID, Account, AccountName, AccountLookup, 
                   SubParent, Parent, VW, Belastingaangifte, 
                   administration, Pattern
            FROM rekeningschema
            WHERE administration = %s AND Account = %s
        """
        
        result = db.execute_query(query, (tenant, account))
        
        if not result:
            return jsonify({'error': 'Account not found'}), 404
        
        return jsonify({
            'success': True,
            'account': result[0]
        })
        
    except Exception as e:
        logger.error(f"Error getting account {account} for tenant {tenant}: {e}")
        return jsonify({'error': 'Failed to get account', 'details': str(e)}), 500



@chart_of_accounts_bp.route('/api/tenant-admin/chart-of-accounts', methods=['POST'])
@cognito_required(required_permissions=[])
def create_account(user_email, user_roles):
    """
    Create new account.
    
    Request Body:
        {
            "account": "1000",
            "accountName": "Kas",
            "accountLookup": "CASH",
            "belastingaangifte": "Activa"
        }
    
    Returns:
        {
            "success": true,
            "account": {...}
        }
    """
    try:
        # Get tenant from request
        tenant = get_current_tenant(request)
        
        # Extract user tenants from JWT
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            jwt_token = auth_header.replace('Bearer ', '').strip()
            user_tenants = get_user_tenants(jwt_token)
        else:
            return jsonify({'error': 'Invalid authorization'}), 401
        
        # Check if user is tenant admin
        if not is_tenant_admin(user_roles, tenant, user_tenants):
            return jsonify({'error': 'Tenant admin access required'}), 403
        
        # Check FIN module access
        if not has_fin_module(tenant):
            return jsonify({'error': 'FIN module not enabled'}), 403
        
        # Get request data
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Request body required'}), 400
        
        # Validate required fields
        try:
            account = validate_account_number(data.get('account', ''))
            account_name = validate_account_name(data.get('accountName', ''))
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        
        # Optional fields
        account_lookup = data.get('accountLookup', '').strip()
        sub_parent = data.get('subParent', '').strip()
        parent = data.get('parent', '').strip()
        vw = data.get('vw', '').strip()
        belastingaangifte = data.get('belastingaangifte', '').strip()
        pattern = 1 if data.get('pattern') else 0
        
        # Initialize database
        test_mode = os.getenv('TEST_MODE', 'false').lower() == 'true'
        db = DatabaseManager(test_mode=test_mode)
        
        # Check if account already exists
        check_query = """
            SELECT 1 FROM rekeningschema
            WHERE administration = %s AND Account = %s
        """
        exists = db.execute_query(check_query, (tenant, account))
        
        if exists:
            return jsonify({'error': 'Account already exists'}), 409
        
        # Insert account
        insert_query = """
            INSERT INTO rekeningschema
            (Account, AccountName, AccountLookup, SubParent, Parent, VW, 
             Belastingaangifte, administration, Pattern)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        db.execute_query(
            insert_query,
            (account, account_name, account_lookup, sub_parent, parent, vw, 
             belastingaangifte, tenant, pattern),
            fetch=False,
            commit=True
        )
        
        # Audit log (TODO: Implement proper audit logging)
        logger.info(f"CREATE_ACCOUNT: {user_email} created account {account} for tenant {tenant}")
        
        return jsonify({
            'success': True,
            'account': {
                'Account': account,
                'AccountName': account_name,
                'AccountLookup': account_lookup,
                'SubParent': sub_parent,
                'Parent': parent,
                'VW': vw,
                'Belastingaangifte': belastingaangifte,
                'administration': tenant,
                'Pattern': pattern
            }
        }), 201
        
    except Exception as e:
        logger.error(f"Error creating account for tenant {tenant}: {e}")
        return jsonify({'error': 'Failed to create account', 'details': str(e)}), 500



@chart_of_accounts_bp.route('/api/tenant-admin/chart-of-accounts/<account>', methods=['PUT'])
@cognito_required(required_permissions=[])
def update_account(user_email, user_roles, account):
    """
    Update existing account.
    
    Note: Account number cannot be changed.
    
    Path Parameters:
        - account: string (required) - Account number
    
    Request Body:
        {
            "accountName": "Updated Name",
            "accountLookup": "UPDATED",
            "belastingaangifte": "Updated Category"
        }
    
    Returns:
        {
            "success": true,
            "account": {...}
        }
    """
    try:
        # Get tenant from request
        tenant = get_current_tenant(request)
        
        # Extract user tenants from JWT
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            jwt_token = auth_header.replace('Bearer ', '').strip()
            user_tenants = get_user_tenants(jwt_token)
        else:
            return jsonify({'error': 'Invalid authorization'}), 401
        
        # Check if user is tenant admin
        if not is_tenant_admin(user_roles, tenant, user_tenants):
            return jsonify({'error': 'Tenant admin access required'}), 403
        
        # Check FIN module access
        if not has_fin_module(tenant):
            return jsonify({'error': 'FIN module not enabled'}), 403
        
        # Get request data
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Request body required'}), 400
        
        # Initialize database
        test_mode = os.getenv('TEST_MODE', 'false').lower() == 'true'
        db = DatabaseManager(test_mode=test_mode)
        
        # Check if account exists and get old values
        check_query = """
            SELECT Account, AccountName, AccountLookup, Belastingaangifte
            FROM rekeningschema
            WHERE administration = %s AND Account = %s
        """
        existing = db.execute_query(check_query, (tenant, account))
        
        if not existing:
            return jsonify({'error': 'Account not found'}), 404
        
        old_values = existing[0]
        
        # Get new values (use old values if not provided)
        account_name = data.get('accountName', old_values['AccountName']).strip()
        account_lookup = data.get('accountLookup', old_values.get('AccountLookup', '')).strip()
        sub_parent = data.get('subParent', old_values.get('SubParent', '')).strip()
        parent = data.get('parent', old_values.get('Parent', '')).strip()
        vw = data.get('vw', old_values.get('VW', '')).strip()
        belastingaangifte = data.get('belastingaangifte', old_values.get('Belastingaangifte', '')).strip()
        pattern = 1 if data.get('pattern', old_values.get('Pattern', 0)) else 0
        
        # Validate account name if provided
        if 'accountName' in data:
            try:
                account_name = validate_account_name(account_name)
            except ValueError as e:
                return jsonify({'error': str(e)}), 400
        
        # Update account
        update_query = """
            UPDATE rekeningschema
            SET AccountName = %s,
                AccountLookup = %s,
                SubParent = %s,
                Parent = %s,
                VW = %s,
                Belastingaangifte = %s,
                Pattern = %s
            WHERE administration = %s AND Account = %s
        """
        
        db.execute_query(
            update_query,
            (account_name, account_lookup, sub_parent, parent, vw, 
             belastingaangifte, pattern, tenant, account),
            fetch=False,
            commit=True
        )
        
        # Audit log (TODO: Implement proper audit logging)
        logger.info(f"UPDATE_ACCOUNT: {user_email} updated account {account} for tenant {tenant}")
        
        return jsonify({
            'success': True,
            'account': {
                'Account': account,
                'AccountName': account_name,
                'AccountLookup': account_lookup,
                'SubParent': sub_parent,
                'Parent': parent,
                'VW': vw,
                'Belastingaangifte': belastingaangifte,
                'administration': tenant,
                'Pattern': pattern
            }
        })
        
    except Exception as e:
        logger.error(f"Error updating account {account} for tenant {tenant}: {e}")
        return jsonify({'error': 'Failed to update account', 'details': str(e)}), 500



@chart_of_accounts_bp.route('/api/tenant-admin/chart-of-accounts/<account>', methods=['DELETE'])
@cognito_required(required_permissions=[])
def delete_account(user_email, user_roles, account):
    """
    Delete account (only if not used in transactions).
    
    Path Parameters:
        - account: string (required) - Account number
    
    Returns:
        {
            "success": true,
            "message": "Account deleted"
        }
        
    Or if account is in use:
        {
            "error": "Cannot delete account that is used in transactions",
            "usage_count": 42
        }
    """
    try:
        # Get tenant from request
        tenant = get_current_tenant(request)
        
        # Extract user tenants from JWT
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            jwt_token = auth_header.replace('Bearer ', '').strip()
            user_tenants = get_user_tenants(jwt_token)
        else:
            return jsonify({'error': 'Invalid authorization'}), 401
        
        # Check if user is tenant admin
        if not is_tenant_admin(user_roles, tenant, user_tenants):
            return jsonify({'error': 'Tenant admin access required'}), 403
        
        # Check FIN module access
        if not has_fin_module(tenant):
            return jsonify({'error': 'FIN module not enabled'}), 403
        
        # Initialize database
        test_mode = os.getenv('TEST_MODE', 'false').lower() == 'true'
        db = DatabaseManager(test_mode=test_mode)
        
        # Check if account exists
        check_query = """
            SELECT 1 FROM rekeningschema
            WHERE administration = %s AND Account = %s
        """
        exists = db.execute_query(check_query, (tenant, account))
        
        if not exists:
            return jsonify({'error': 'Account not found'}), 404
        
        # Check if account is used in transactions
        usage_count = is_account_used_in_transactions(tenant, account)
        
        if usage_count > 0:
            return jsonify({
                'error': 'Cannot delete account that is used in transactions',
                'usage_count': usage_count
            }), 409
        
        # Delete account
        delete_query = """
            DELETE FROM rekeningschema
            WHERE administration = %s AND Account = %s
        """
        
        db.execute_query(delete_query, (tenant, account), fetch=False, commit=True)
        
        # Audit log (TODO: Implement proper audit logging)
        logger.info(f"DELETE_ACCOUNT: {user_email} deleted account {account} for tenant {tenant}")
        
        return jsonify({
            'success': True,
            'message': 'Account deleted'
        })
        
    except Exception as e:
        logger.error(f"Error deleting account {account} for tenant {tenant}: {e}")
        return jsonify({'error': 'Failed to delete account', 'details': str(e)}), 500



@chart_of_accounts_bp.route('/api/tenant-admin/chart-of-accounts/export', methods=['GET'])
@cognito_required(required_permissions=[])
def export_accounts(user_email, user_roles):
    """
    Export all accounts to Excel file.
    
    Returns:
        Excel file download with all accounts
    """
    try:
        # Get tenant from request
        tenant = get_current_tenant(request)
        
        # Extract user tenants from JWT
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            jwt_token = auth_header.replace('Bearer ', '').strip()
            user_tenants = get_user_tenants(jwt_token)
        else:
            return jsonify({'error': 'Invalid authorization'}), 401
        
        # Check if user is tenant admin
        if not is_tenant_admin(user_roles, tenant, user_tenants):
            return jsonify({'error': 'Tenant admin access required'}), 403
        
        # Check FIN module access
        if not has_fin_module(tenant):
            return jsonify({'error': 'FIN module not enabled'}), 403
        
        # Initialize database
        test_mode = os.getenv('TEST_MODE', 'false').lower() == 'true'
        db = DatabaseManager(test_mode=test_mode)
        
        # Get all accounts
        query = """
            SELECT AccountID, Account, AccountName, AccountLookup, 
                   SubParent, Parent, VW, Belastingaangifte, 
                   administration, Pattern
            FROM rekeningschema
            WHERE administration = %s
            ORDER BY Account
        """
        
        accounts = db.execute_query(query, (tenant,))
        
        # Create Excel file
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chart of Accounts"
        
        # Headers
        ws.append(['Account', 'AccountName', 'AccountLookup', 'SubParent', 'Parent', 
                   'VW', 'Belastingaangifte', 'Pattern'])
        
        # Data
        for account in accounts:
            ws.append([
                account['Account'],
                account['AccountName'],
                account.get('AccountLookup', ''),
                account.get('SubParent', ''),
                account.get('Parent', ''),
                account.get('VW', ''),
                account.get('Belastingaangifte', ''),
                1 if account.get('Pattern') else 0
            ])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Audit log (TODO: Implement proper audit logging)
        logger.info(f"EXPORT_ACCOUNTS: {user_email} exported {len(accounts)} accounts for tenant {tenant}")
        
        # Generate filename with tenant and date
        filename = f'chart_of_accounts_{tenant}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error exporting accounts for tenant {tenant}: {e}")
        return jsonify({'error': 'Failed to export accounts', 'details': str(e)}), 500



@chart_of_accounts_bp.route('/api/tenant-admin/chart-of-accounts/import', methods=['POST'])
@cognito_required(required_permissions=[])
def import_accounts(user_email, user_roles):
    """
    Import accounts from Excel file.
    
    Request:
        - file: Excel file upload (.xlsx or .xls)
    
    Returns:
        {
            "success": true,
            "imported": 10,
            "updated": 5,
            "total": 15
        }
    """
    try:
        # Get tenant from request
        tenant = get_current_tenant(request)
        
        # Extract user tenants from JWT
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            jwt_token = auth_header.replace('Bearer ', '').strip()
            user_tenants = get_user_tenants(jwt_token)
        else:
            return jsonify({'error': 'Invalid authorization'}), 401
        
        # Check if user is tenant admin
        if not is_tenant_admin(user_roles, tenant, user_tenants):
            return jsonify({'error': 'Tenant admin access required'}), 403
        
        # Check FIN module access
        if not has_fin_module(tenant):
            return jsonify({'error': 'FIN module not enabled'}), 403
        
        # Get uploaded file
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        
        if not file.filename:
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file type. Must be Excel file (.xlsx or .xls)'}), 400
        
        # Parse Excel
        import openpyxl
        
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            return jsonify({'error': 'Failed to parse Excel file', 'details': str(e)}), 400
        
        # Validate headers
        headers = [cell.value for cell in ws[1]]
        expected = ['Account', 'AccountName', 'AccountLookup', 'SubParent', 'Parent', 
                   'VW', 'Belastingaangifte', 'Pattern']
        
        if headers != expected:
            return jsonify({
                'error': 'Invalid Excel format',
                'expected_headers': expected,
                'found_headers': headers
            }), 400
        
        # Parse rows
        accounts_to_import = []
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
                
            account, name, lookup, sub_parent, parent, vw, tax, pattern = row
            
            # Validate
            if not account:
                errors.append(f"Row {row_num}: Account number required")
                continue
            if not name:
                errors.append(f"Row {row_num}: Account name required")
                continue
            
            accounts_to_import.append({
                'account': str(account).strip(),
                'name': str(name).strip(),
                'lookup': str(lookup).strip() if lookup else '',
                'sub_parent': str(sub_parent).strip() if sub_parent else '',
                'parent': str(parent).strip() if parent else '',
                'vw': str(vw).strip() if vw else '',
                'tax': str(tax).strip() if tax else '',
                'pattern': 1 if pattern else 0
            })
        
        if errors:
            return jsonify({
                'success': False,
                'errors': errors,
                'parsed': len(accounts_to_import)
            }), 400
        
        # Initialize database
        test_mode = os.getenv('TEST_MODE', 'false').lower() == 'true'
        db = DatabaseManager(test_mode=test_mode)
        
        # Import accounts (upsert)
        imported = 0
        updated = 0
        
        for acc in accounts_to_import:
            # Check if exists
            check_query = """
                SELECT 1 FROM rekeningschema
                WHERE administration = %s AND Account = %s
            """
            exists = db.execute_query(check_query, (tenant, acc['account']))
            
            if exists:
                # Update
                update_query = """
                    UPDATE rekeningschema
                    SET AccountName = %s, AccountLookup = %s, SubParent = %s, 
                        Parent = %s, VW = %s, Belastingaangifte = %s, Pattern = %s
                    WHERE administration = %s AND Account = %s
                """
                db.execute_query(
                    update_query,
                    (acc['name'], acc['lookup'], acc['sub_parent'], acc['parent'], 
                     acc['vw'], acc['tax'], acc['pattern'], tenant, acc['account']),
                    fetch=False,
                    commit=True
                )
                updated += 1
            else:
                # Insert
                insert_query = """
                    INSERT INTO rekeningschema
                    (Account, AccountName, AccountLookup, SubParent, Parent, VW, 
                     Belastingaangifte, administration, Pattern)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                db.execute_query(
                    insert_query,
                    (acc['account'], acc['name'], acc['lookup'], acc['sub_parent'], 
                     acc['parent'], acc['vw'], acc['tax'], tenant, acc['pattern']),
                    fetch=False,
                    commit=True
                )
                imported += 1
        
        # Audit log (TODO: Implement proper audit logging)
        logger.info(f"IMPORT_ACCOUNTS: {user_email} imported {imported} and updated {updated} accounts for tenant {tenant}")
        
        return jsonify({
            'success': True,
            'imported': imported,
            'updated': updated,
            'total': imported + updated
        })
        
    except Exception as e:
        logger.error(f"Error importing accounts for tenant {tenant}: {e}")
        return jsonify({'error': 'Failed to import accounts', 'details': str(e)}), 500

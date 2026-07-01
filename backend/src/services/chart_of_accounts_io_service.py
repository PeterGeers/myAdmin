"""
Chart of Accounts Import/Export Service

Handles Excel import and export operations for chart of accounts.
Extracted from chart_of_accounts_routes.py for file size management.
"""

import json
import logging
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Tuple

from database import DatabaseManager
from dialect_helpers import dialect

logger = logging.getLogger(__name__)

# Expected headers in import/export Excel files
EXCEL_HEADERS = [
    "Account",
    "AccountName",
    "AccountLookup",
    "SubParent",
    "Parent",
    "VW",
    "Belastingaangifte",
    "Pattern",
]


class ChartOfAccountsIOService:
    """Service for importing and exporting chart of accounts data."""

    def __init__(self, db: DatabaseManager) -> None:
        self.db = db

    def export_to_excel(self, tenant: str) -> Tuple[BytesIO, str, int]:
        """
        Export all accounts for a tenant to an Excel BytesIO stream.

        Args:
            tenant: Tenant identifier.

        Returns:
            Tuple of (BytesIO stream, filename, account_count).
        """
        import openpyxl

        query = f"""
            SELECT AccountID, Account, AccountName, AccountLookup,
                   SubParent, Parent, VW, Belastingaangifte,
                   administration,
                   {dialect.ifnull(dialect.json_extract("parameters", "$.bank_account"), "false")} as bank_account,
                   {dialect.json_unquote_extract("parameters", "$.iban")} as iban
            FROM rekeningschema
            WHERE administration = %s
            ORDER BY Account
        """

        accounts = self.db.execute_query(query, (tenant,))
        accounts = accounts if accounts else []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chart of Accounts"

        # Headers
        ws.append(EXCEL_HEADERS)

        # Data rows
        for account in accounts:
            ws.append(
                [
                    account["Account"],
                    account["AccountName"],
                    account.get("AccountLookup", ""),
                    account.get("SubParent", ""),
                    account.get("Parent", ""),
                    account.get("VW", ""),
                    account.get("Belastingaangifte", ""),
                    1 if account.get("bank_account") else 0,
                ]
            )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"chart_of_accounts_{tenant}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        return output, filename, len(accounts)

    def import_from_excel(self, tenant: str, file_stream) -> Dict[str, Any]:
        """
        Import accounts from an Excel file stream (upsert logic).

        Args:
            tenant: Tenant identifier.
            file_stream: File-like object containing the Excel data.

        Returns:
            Dict with 'success', 'imported', 'updated', 'total' or 'errors'.
        """
        import openpyxl

        try:
            wb = openpyxl.load_workbook(file_stream)
            ws = wb.active
        except Exception as e:
            return {
                "success": False,
                "error": "Failed to parse Excel file",
                "details": str(e),
            }

        # Validate headers
        headers = [cell.value for cell in ws[1]]
        if headers != EXCEL_HEADERS:
            return {
                "success": False,
                "error": "Invalid Excel format",
                "expected_headers": EXCEL_HEADERS,
                "found_headers": headers,
            }

        # Parse rows
        accounts_to_import: List[Dict[str, Any]] = []
        errors: List[str] = []

        for row_num, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(row):  # Skip empty rows
                continue

            account, name, lookup, sub_parent, parent, vw, tax, pattern = row

            if not account:
                errors.append(f"Row {row_num}: Account number required")
                continue
            if not name:
                errors.append(f"Row {row_num}: Account name required")
                continue

            accounts_to_import.append(
                {
                    "account": str(account).strip(),
                    "name": str(name).strip(),
                    "lookup": str(lookup).strip() if lookup else "",
                    "sub_parent": str(sub_parent).strip() if sub_parent else "",
                    "parent": str(parent).strip() if parent else "",
                    "vw": str(vw).strip() if vw else "",
                    "tax": str(tax).strip() if tax else "",
                    "bank_account": bool(pattern),
                    "iban": str(lookup).strip() if lookup and pattern else None,
                }
            )

        if errors:
            return {
                "success": False,
                "errors": errors,
                "parsed": len(accounts_to_import),
            }

        # Upsert accounts
        imported = 0
        updated = 0

        for acc in accounts_to_import:
            params_dict: Dict[str, Any] = {}
            if acc["bank_account"]:
                params_dict["bank_account"] = True
            if acc["iban"]:
                params_dict["iban"] = acc["iban"]
            parameters_json = json.dumps(params_dict) if params_dict else None

            exists = self.db.execute_query(
                "SELECT 1 FROM rekeningschema WHERE administration = %s AND Account = %s",
                (tenant, acc["account"]),
            )

            if exists:
                self.db.execute_query(
                    """
                    UPDATE rekeningschema
                    SET AccountName = %s, AccountLookup = %s, SubParent = %s,
                        Parent = %s, VW = %s, Belastingaangifte = %s,
                        parameters = %s
                    WHERE administration = %s AND Account = %s
                    """,
                    (
                        acc["name"],
                        acc["lookup"],
                        acc["sub_parent"],
                        acc["parent"],
                        acc["vw"],
                        acc["tax"],
                        parameters_json,
                        tenant,
                        acc["account"],
                    ),
                    fetch=False,
                    commit=True,
                )
                updated += 1
            else:
                self.db.execute_query(
                    """
                    INSERT INTO rekeningschema
                    (Account, AccountName, AccountLookup, SubParent, Parent, VW,
                     Belastingaangifte, administration, parameters)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        acc["account"],
                        acc["name"],
                        acc["lookup"],
                        acc["sub_parent"],
                        acc["parent"],
                        acc["vw"],
                        acc["tax"],
                        tenant,
                        parameters_json,
                    ),
                    fetch=False,
                    commit=True,
                )
                imported += 1

        return {
            "success": True,
            "imported": imported,
            "updated": updated,
            "total": imported + updated,
        }

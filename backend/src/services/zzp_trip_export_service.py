"""
TripExportService: PDF/CSV/XLSX export generation.

Generates export files for the Rittenregistratie module.
Supports PDF (Belastingdienst-compliant), CSV, and XLSX formats.

Business rules:
- Exports are scoped by tenant, vehicle, and year
- Optional filters (trip_category, contact_id, date range)
- PDF format must be Belastingdienst-compliant (per vehicle, per year, all required fields)
- Yearly totals included at bottom (zakelijk/privé/woon-werk)
- CSV uses standard columns: Datum, Vertrekadres, Bestemming, Begin KM, Eind KM, Afstand, Categorie, Doel, Klant

Reference: .kiro/specs/ZZP/rittenregistratie/design.md §4.5
"""

import io
import logging
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, numbers

from db_exceptions import DatabaseError
from dialect_helpers import dialect

logger = logging.getLogger(__name__)


class TripExportService:
    """PDF/CSV/XLSX export generation."""

    def __init__(self, db, parameter_service=None):
        """Initialize TripExportService.

        Args:
            db: DatabaseManager instance for database access.
            parameter_service: Optional ParameterService for tenant-scoped config.
        """
        self.db = db
        self.parameter_service = parameter_service

    def export_pdf(self, tenant: str, vehicle_id: int, year: int, filters: dict = None) -> bytes:
        """Generate a Belastingdienst-compliant PDF export.

        Includes all trips for the vehicle/year with yearly totals
        broken down by category (zakelijk/privé/woon-werk).

        Args:
            tenant: Administration/tenant identifier.
            vehicle_id: Vehicle to export trips for.
            year: Calendar year to export.
            filters: Optional additional filters (trip_category, contact_id, etc.).

        Returns:
            PDF file content as bytes.
        """
        trips = self._get_trips_for_export(tenant, vehicle_id, year, filters)
        vehicle = self._get_vehicle_info(tenant, vehicle_id)

        # Calculate yearly totals by category
        totals = self._calculate_totals(trips)

        # Render HTML from template
        html = self._render_trip_pdf_html(vehicle, year, trips, totals)

        # Convert HTML to PDF via weasyprint
        pdf_bytes = self._html_to_pdf(html)
        return pdf_bytes

    def export_csv(self, tenant: str, vehicle_id: int, year: int, filters: dict = None) -> bytes:
        """Generate a CSV export of trips.

        Columns: Datum, Vertrekadres, Bestemming, Begin KM, Eind KM,
        Afstand, Categorie, Doel, Klant.

        Uses semicolon separator (Dutch standard) and UTF-8 with BOM
        for Excel compatibility.

        Args:
            tenant: Administration/tenant identifier.
            vehicle_id: Vehicle to export trips for.
            year: Calendar year to export.
            filters: Optional additional filters (trip_category, contact_id, etc.).

        Returns:
            CSV file content as bytes (UTF-8-BOM encoded).
        """
        trips = self._get_trips_for_export(tenant, vehicle_id, year, filters)

        # Map trip data to Dutch column names
        column_mapping = {
            "trip_date": "Datum",
            "start_address": "Vertrekadres",
            "end_address": "Bestemming",
            "start_odometer": "Begin KM",
            "end_odometer": "Eind KM",
            "distance_km": "Afstand",
            "trip_category": "Categorie",
            "trip_purpose": "Doel",
            "contact_name": "Klant",
        }

        # Build rows with only the export columns
        rows = []
        for trip in trips:
            rows.append({
                dutch_name: trip.get(field_name, "")
                for field_name, dutch_name in column_mapping.items()
            })

        df = pd.DataFrame(rows, columns=list(column_mapping.values()))

        # Export to CSV: semicolon separator (Dutch standard), UTF-8 with BOM
        buffer = io.StringIO()
        df.to_csv(buffer, index=False, sep=";", quoting=1)  # quoting=1 = csv.QUOTE_ALL

        csv_content = buffer.getvalue()
        # Encode with UTF-8 BOM for Excel compatibility
        return b"\xef\xbb\xbf" + csv_content.encode("utf-8")

    def export_xlsx(self, tenant: str, vehicle_id: int, year: int, filters: dict = None) -> bytes:
        """Generate an XLSX export with styled headers.

        Same data as CSV but with openpyxl formatting (bold headers,
        column widths, number formatting for odometer/distance).

        Args:
            tenant: Administration/tenant identifier.
            vehicle_id: Vehicle to export trips for.
            year: Calendar year to export.
            filters: Optional additional filters (trip_category, contact_id, etc.).

        Returns:
            XLSX file content as bytes.
        """
        trips = self._get_trips_for_export(tenant, vehicle_id, year, filters)

        # Column mapping: (internal field, Dutch header, column width)
        columns = [
            ("trip_date", "Datum", 12),
            ("start_address", "Vertrekadres", 35),
            ("end_address", "Bestemming", 35),
            ("start_odometer", "Begin KM", 12),
            ("end_odometer", "Eind KM", 12),
            ("distance_km", "Afstand", 12),
            ("trip_category", "Categorie", 15),
            ("trip_purpose", "Doel", 20),
            ("contact_name", "Klant", 25),
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = f"Rittenregistratie {year}"

        bold_font = Font(bold=True)

        # Write header row with bold styling
        for col_idx, (_, header, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font

        # Set column widths
        for col_idx, (_, _, width) in enumerate(columns, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = width

        # Identify numeric columns for number formatting (0-indexed within columns list)
        numeric_fields = {"start_odometer", "end_odometer", "distance_km"}

        # Write data rows
        for row_idx, trip in enumerate(trips, start=2):
            for col_idx, (field, _, _) in enumerate(columns, start=1):
                value = trip.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if field in numeric_fields and value != "":
                    cell.number_format = numbers.FORMAT_NUMBER

        # Save workbook to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def get_yearly_summary(self, tenant: str, vehicle_id: int, year: int) -> dict:
        """Get aggregated yearly summary data for reports/exports.

        Returns totals per category and monthly breakdown suitable
        for inclusion in export footers or standalone summary reports.

        Args:
            tenant: Administration/tenant identifier.
            vehicle_id: Vehicle to summarize.
            year: Calendar year to summarize.

        Returns:
            Dict with keys: year, vehicle_id, total_km, zakelijk_km,
            prive_km, woonwerk_km, trip_count, monthly_breakdown.
        """
        trips = self._get_trips_for_export(tenant, vehicle_id, year)
        totals = self._calculate_totals(trips)

        # Build monthly breakdown: group trips by month, sum distance per category
        monthly_breakdown = self._build_monthly_breakdown(trips)

        return {
            "year": year,
            "vehicle_id": vehicle_id,
            "total_km": totals["total_km"],
            "zakelijk_km": totals["zakelijk_km"],
            "prive_km": totals["prive_km"],
            "woonwerk_km": totals["woonwerk_km"],
            "trip_count": totals["trip_count"],
            "monthly_breakdown": monthly_breakdown,
        }

    @staticmethod
    def _build_monthly_breakdown(trips: List[dict]) -> List[dict]:
        """Group trips by month and sum distance per category.

        Args:
            trips: List of formatted trip dicts (with trip_date as string "YYYY-MM-DD").

        Returns:
            List of dicts sorted by month, each with keys:
            month (str "YYYY-MM"), zakelijk (int), prive (int), woonwerk (int).
        """
        monthly: dict = {}

        for trip in trips:
            trip_date = trip.get("trip_date", "")
            if not trip_date or len(trip_date) < 7:
                continue

            # Extract "YYYY-MM" from the date string
            month_key = trip_date[:7]

            if month_key not in monthly:
                monthly[month_key] = {"zakelijk": 0, "prive": 0, "woonwerk": 0}

            distance = trip.get("distance_km") or 0
            category = (trip.get("trip_category") or "").lower()

            if category == "zakelijk":
                monthly[month_key]["zakelijk"] += distance
            elif category in ("privé", "prive"):
                monthly[month_key]["prive"] += distance
            elif category in ("woon-werk", "woonwerk"):
                monthly[month_key]["woonwerk"] += distance

        # Return sorted by month
        return [
            {"month": month_key, **values}
            for month_key, values in sorted(monthly.items())
        ]

    def _get_vehicle_info(self, tenant: str, vehicle_id: int) -> dict:
        """Load vehicle details for the PDF header.

        Args:
            tenant: Administration/tenant identifier.
            vehicle_id: Vehicle to look up.

        Returns:
            Dict with vehicle info (license_plate, make, model, etc.)
            or a minimal fallback dict if not found.
        """
        try:
            rows = self.db.execute_query(
                "SELECT id, license_plate, make, model, year_built, "
                "vehicle_type, odometer_unit "
                "FROM zzp_vehicles "
                "WHERE id = %s AND administration = %s",
                (int(vehicle_id), tenant),
            )
            if rows and len(rows) > 0:
                return dict(rows[0])
        except DatabaseError:
            logger.warning(
                "Failed to load vehicle info: tenant=%s, vehicle_id=%s",
                tenant, vehicle_id,
            )
        return {"id": vehicle_id, "license_plate": "Onbekend", "make": "", "model": ""}

    @staticmethod
    def _calculate_totals(trips: List[dict]) -> dict:
        """Calculate yearly totals by category.

        Args:
            trips: List of formatted trip dicts.

        Returns:
            Dict with total_km, zakelijk_km, prive_km, woonwerk_km, trip_count.
        """
        total_km = 0
        zakelijk_km = 0
        prive_km = 0
        woonwerk_km = 0

        for trip in trips:
            distance = trip.get("distance_km") or 0
            total_km += distance
            category = (trip.get("trip_category") or "").lower()
            if category == "zakelijk":
                zakelijk_km += distance
            elif category in ("privé", "prive"):
                prive_km += distance
            elif category in ("woon-werk", "woonwerk"):
                woonwerk_km += distance

        return {
            "total_km": total_km,
            "zakelijk_km": zakelijk_km,
            "prive_km": prive_km,
            "woonwerk_km": woonwerk_km,
            "trip_count": len(trips),
        }

    def _render_trip_pdf_html(
        self, vehicle: dict, year: int, trips: List[dict], totals: dict
    ) -> str:
        """Render the Belastingdienst-compliant HTML for trip export.

        Args:
            vehicle: Vehicle info dict.
            year: Calendar year.
            trips: List of formatted trip dicts.
            totals: Calculated totals dict.

        Returns:
            Complete HTML string ready for PDF conversion.
        """
        license_plate = vehicle.get("license_plate", "Onbekend")
        make = vehicle.get("make") or ""
        model = vehicle.get("model") or ""
        vehicle_desc = f"{make} {model}".strip() or license_plate

        # Build table rows
        rows_html = ""
        for trip in trips:
            trip_date = trip.get("trip_date", "")
            start_address = trip.get("start_address", "")
            end_address = trip.get("end_address", "")
            start_km = trip.get("start_odometer", "")
            end_km = trip.get("end_odometer", "")
            distance = trip.get("distance_km", "")
            category = trip.get("trip_category", "")
            purpose = trip.get("trip_purpose", "")
            client = trip.get("contact_name", "") or ""

            rows_html += (
                f"<tr>"
                f"<td>{trip_date}</td>"
                f"<td>{start_address}</td>"
                f"<td>{end_address}</td>"
                f'<td class="num">{start_km}</td>'
                f'<td class="num">{end_km}</td>'
                f'<td class="num">{distance}</td>'
                f"<td>{category}</td>"
                f"<td>{purpose}</td>"
                f"<td>{client}</td>"
                f"</tr>"
            )

        html = _TRIP_PDF_TEMPLATE.replace("{{year}}", str(year))
        html = html.replace("{{license_plate}}", license_plate)
        html = html.replace("{{vehicle_desc}}", vehicle_desc)
        html = html.replace("{{trip_rows}}", rows_html)
        html = html.replace("{{trip_count}}", str(totals["trip_count"]))
        html = html.replace("{{total_km}}", str(totals["total_km"]))
        html = html.replace("{{zakelijk_km}}", str(totals["zakelijk_km"]))
        html = html.replace("{{prive_km}}", str(totals["prive_km"]))
        html = html.replace("{{woonwerk_km}}", str(totals["woonwerk_km"]))

        return html

    @staticmethod
    def _html_to_pdf(html: str) -> bytes:
        """Convert HTML string to PDF bytes via weasyprint.

        Args:
            html: Rendered HTML string.

        Returns:
            PDF content as bytes.

        Raises:
            RuntimeError: If weasyprint is not installed.
        """
        try:
            import weasyprint
        except ImportError:
            logger.error("weasyprint not installed — cannot generate PDF")
            raise RuntimeError(
                "PDF generation requires weasyprint. "
                "Install with: pip install weasyprint>=60.0"
            )

        return weasyprint.HTML(string=html).write_pdf()

    def _get_trips_for_export(
        self, tenant: str, vehicle_id: int, year: int, filters: dict = None
    ) -> List[dict]:
        """Query trips from the database for export.

        Retrieves all non-cancelled trips for the given tenant, vehicle,
        and year. Applies optional filters for further narrowing results.
        Results are ordered by trip_date ASC, start_odometer ASC for
        chronological export output.

        Args:
            tenant: Administration/tenant identifier.
            vehicle_id: Vehicle to query trips for.
            year: Calendar year to filter on.
            filters: Optional dict with additional filters:
                - trip_category: Filter by category (Zakelijk, Privé, Woon-werk)
                - contact_id: Filter by linked client
                - date_from: Start date (within the year)
                - date_to: End date (within the year)

        Returns:
            List of trip dicts ordered by date ascending.

        Raises:
            DatabaseError: On database query failure.
        """
        filters = filters or {}

        where_clauses = [
            "t.administration = %s",
            "t.vehicle_id = %s",
            f"{dialect.year('t.trip_date')} = %s",
            "t.is_cancelled = FALSE",
        ]
        params: list = [tenant, int(vehicle_id), int(year)]

        # Apply optional filters
        if filters.get("trip_category"):
            where_clauses.append("t.trip_category = %s")
            params.append(filters["trip_category"])

        if filters.get("contact_id"):
            where_clauses.append("t.contact_id = %s")
            params.append(int(filters["contact_id"]))

        if filters.get("date_from"):
            where_clauses.append("t.trip_date >= %s")
            params.append(filters["date_from"])

        if filters.get("date_to"):
            where_clauses.append("t.trip_date <= %s")
            params.append(filters["date_to"])

        where_sql = " AND ".join(where_clauses)

        query = (
            "SELECT t.id, t.trip_date, t.start_time, t.end_time, "
            "t.start_address, t.end_address, "
            "t.start_odometer, t.end_odometer, t.distance_km, "
            "t.trip_category, t.trip_purpose, "
            "t.route_description, t.contact_id, t.project_name, "
            "t.notes, t.is_billable, t.is_billed, t.is_gap_fill, "
            "c.company_name as contact_name "
            "FROM zzp_trips t "
            "LEFT JOIN contacts c ON t.contact_id = c.id AND c.administration = t.administration "
            f"WHERE {where_sql} "
            "ORDER BY t.trip_date ASC, t.start_odometer ASC"
        )

        try:
            rows = self.db.execute_query(query, tuple(params)) or []
            return [self._format_export_row(row) for row in rows]
        except DatabaseError:
            logger.error(
                "Failed to query trips for export: tenant=%s, vehicle=%s, year=%s",
                tenant, vehicle_id, year,
            )
            raise

    @staticmethod
    def _format_export_row(row: dict) -> dict:
        """Format a raw database row for export output.

        Converts date objects to ISO strings, ensures integer types
        for odometer/distance fields.

        Args:
            row: Raw database row dict.

        Returns:
            Formatted row dict ready for export processing.
        """
        row = dict(row)

        # Convert date/time fields to strings
        for key in ("trip_date", "start_time", "end_time"):
            val = row.get(key)
            if val is not None and hasattr(val, "isoformat") and not isinstance(val, str):
                row[key] = val.isoformat()

        # Ensure integer odometer values
        for key in ("start_odometer", "end_odometer", "distance_km"):
            if key in row and row[key] is not None:
                row[key] = int(row[key])

        # Ensure boolean fields
        for key in ("is_billable", "is_billed", "is_gap_fill"):
            if key in row:
                row[key] = bool(row[key])

        return row


# ---------------------------------------------------------------------------
# Belastingdienst-compliant trip export HTML template
# ---------------------------------------------------------------------------

_TRIP_PDF_TEMPLATE = """<!DOCTYPE html>
<html lang="nl">
<head>
<meta charset="utf-8"/>
<style>
  @page { size: A4 landscape; margin: 1.5cm; }
  body { font-family: Arial, sans-serif; font-size: 8pt; color: #333; margin: 0; }
  h1 { font-size: 14pt; margin: 0 0 4px 0; }
  .subtitle { font-size: 9pt; color: #555; margin-bottom: 12px; }
  table.trips { width: 100%; border-collapse: collapse; margin-top: 8px; }
  table.trips th {
    background: #f0f0f0; padding: 4px 6px; text-align: left;
    border-bottom: 2px solid #999; font-size: 7.5pt; color: #444;
  }
  table.trips td { padding: 3px 6px; border-bottom: 1px solid #ddd; font-size: 7.5pt; }
  table.trips tr:nth-child(even) { background: #fafafa; }
  .num { text-align: right; }
  .totals { margin-top: 16px; page-break-inside: avoid; }
  .totals h2 { font-size: 10pt; margin: 0 0 6px 0; }
  table.summary { border-collapse: collapse; }
  table.summary td { padding: 3px 12px 3px 0; font-size: 9pt; }
  table.summary .label { color: #555; }
  table.summary .value { font-weight: bold; text-align: right; }
  .footer { margin-top: 20px; font-size: 7pt; color: #999;
    border-top: 1px solid #ccc; padding-top: 6px; }
</style>
</head>
<body>
<h1>Rittenregistratie {{year}}</h1>
<div class="subtitle">
  Voertuig: {{vehicle_desc}} &mdash; Kenteken: {{license_plate}}
</div>

<table class="trips">
<thead>
<tr>
  <th>Datum</th>
  <th>Vertrekadres</th>
  <th>Bestemming</th>
  <th class="num">Begin KM</th>
  <th class="num">Eind KM</th>
  <th class="num">Afstand</th>
  <th>Categorie</th>
  <th>Doel</th>
  <th>Klant</th>
</tr>
</thead>
<tbody>
{{trip_rows}}
</tbody>
</table>

<div class="totals">
<h2>Jaaroverzicht {{year}}</h2>
<table class="summary">
  <tr><td class="label">Aantal ritten:</td><td class="value">{{trip_count}}</td></tr>
  <tr><td class="label">Totaal kilometers:</td><td class="value">{{total_km}} km</td></tr>
  <tr><td class="label">Zakelijk:</td><td class="value">{{zakelijk_km}} km</td></tr>
  <tr><td class="label">Priv\u00e9:</td><td class="value">{{prive_km}} km</td></tr>
  <tr><td class="label">Woon-werk:</td><td class="value">{{woonwerk_km}} km</td></tr>
</table>
</div>

<div class="footer">
  Gegenereerd t.b.v. Belastingdienst &mdash; Rittenregistratie {{year}} &mdash; Kenteken {{license_plate}}
</div>
</body>
</html>"""
